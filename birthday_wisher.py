from datetime import datetime
from twilio.rest import Client
from http import server
import imghdr
from email.message import EmailMessage
from openpyxl import workbook, load_workbook
import smtplib
import os


def send_email(sender_email, sender_passwd, recipient_name, recipient_email, message):
    DOMAIN = 'smtp.gmail.com'
    PORT = 587
    msg = EmailMessage()

    msg['Subject'] = "Birthday wish"
    msg['From'] = sender_email
    msg['To'] = recipient_email
    msg.set_content("Image attached!")

    server = smtplib.SMTP(DOMAIN, PORT)
    try:
        with open('birthdaycard.jpg', 'rb') as f:
            file_data = f.read()
            file_type = imghdr.what(f.name)
            file_name = f.name

        msg.add_attachment(file_data, maintype='image',
                           subtype=file_type, filename=file_name)
        print(f"Sending your wishes to {recipient_name}...")
        server.starttls()
        server.login(sender_email, sender_passwd)
        server.send_message(msg)

        print('Mail sent!')

    except Exception as err:
        print(err)

    finally:
        server.quit()


def format_date(date):
    date_timestamp = date.split(' ')
    date = date_timestamp[0]
    return date.split('-')


account_sid = os.getenv("ACCOUNT_SID")
auth_token = os.getenv("AUTH_TOKEN")
from_ = os.getenv("FROM")
whatsapp_no = os.getenv("WHATSAPP_NO")

client = Client(account_sid, auth_token)


my_email = input("Sender email: ")
my_passwd = input("Sender password: ")

wb = load_workbook('birthdays.xlsx')
ws = wb.active
max_col = ws.max_column
max_row = ws.max_row

present_day = datetime.now().day
present_month = datetime.now().month

for i in range(2, max_row + 1):
    reciever_name = ws.cell(row=i, column=1).value
    current_date = str(ws.cell(row=i, column=2).value)
    reciever_email = ws.cell(row=i, column=3).value
    mob_number = ws.cell(row=i, column=4).value
    birthday_message = f"Happy Birthday,{reciever_name}!!"

    year, month, day = format_date(current_date)
    is_birthday = int(day) == present_day and int(month) == present_month

    if is_birthday:
        send_email(my_email, my_passwd, reciever_name,
                   reciever_email, birthday_message)
